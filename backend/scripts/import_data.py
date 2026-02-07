"""Import data from SupportMind Excel file into the database."""

import asyncio
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
from sqlalchemy import func, select, text

from vector_db.database import Base, async_session_maker, engine
from vector_db.models.conversation import Conversation
from vector_db.models.kb_lineage import KBLineage
from vector_db.models.knowledge_article import KnowledgeArticle
from vector_db.models.learning_event import LearningEvent
from vector_db.models.placeholder import Placeholder
from vector_db.models.question import Question
from vector_db.models.script import Script
from vector_db.models.ticket import Ticket

DATA_FILE = Path(__file__).resolve().parent.parent.parent / "data" / "SupportMind__Final_Data.xlsx"

BATCH_SIZE = 500


def parse_datetime(val: str | datetime | None) -> datetime | None:
    """Parse a datetime value from Excel (may be string or datetime object)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=UTC) if val.tzinfo is None else val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d %H:%M:%S").replace(tzinfo=UTC)
    except (ValueError, TypeError):
        return None


def read_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[dict]:
    """Read a sheet and return a list of dicts (one per non-empty row)."""
    ws = wb[sheet_name]
    headers = [str(h).strip() for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        # Pad short rows with None
        padded = list(row) + [None] * (len(headers) - len(row))
        rows.append(dict(zip(headers, padded, strict=True)))
    return rows


def str_or_none(val: object) -> str | None:
    """Convert a value to string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


async def import_all() -> None:
    """Import all data from the Excel file into the database."""
    if not DATA_FILE.exists():
        print(f"ERROR: Data file not found at {DATA_FILE}")
        return

    # Ensure pgvector extension and tables exist
    async with engine.begin() as conn:
        await conn.execute(text("CREATE EXTENSION IF NOT EXISTS vector"))
        await conn.run_sync(Base.metadata.create_all)

    print(f"Reading {DATA_FILE.name}...")
    wb = openpyxl.load_workbook(DATA_FILE, read_only=True)

    async with async_session_maker() as session:
        # Check if data already exists (idempotent)
        result = await session.execute(select(func.count()).select_from(KnowledgeArticle))
        if result.scalar() > 0:
            print("Data already imported. Skipping. (Drop tables to re-import.)")
            wb.close()
            return

        # 1. Placeholders
        rows = read_sheet(wb, "Placeholder_Dictionary")
        placeholders = [
            Placeholder(
                name=str(r["Placeholder"]).strip(),
                description=str_or_none(r.get("Meaning")),
                default_value=str_or_none(r.get("Example")),
            )
            for r in rows
        ]
        session.add_all(placeholders)
        await session.flush()
        print(f"  Placeholders: {len(placeholders)}")

        # 2. Knowledge Articles
        rows = read_sheet(wb, "Knowledge_Articles")
        for i in range(0, len(rows), BATCH_SIZE):
            batch = rows[i : i + BATCH_SIZE]
            articles = [
                KnowledgeArticle(
                    id=str(r["KB_Article_ID"]).strip(),
                    title=str(r["Title"]).strip(),
                    body=str(r["Body"]).strip(),
                    tags=str_or_none(r.get("Tags")),
                    module=str_or_none(r.get("Module")),
                    category=str_or_none(r.get("Category")),
                    status=str(r.get("Status", "active")).strip().lower(),
                    source_type=str_or_none(r.get("Source_Type")),
                )
                for r in batch
            ]
            session.add_all(articles)
            await session.flush()
        print(f"  Knowledge Articles: {len(rows)}")

        # 3. Scripts
        rows = read_sheet(wb, "Scripts_Master")
        for i in range(0, len(rows), BATCH_SIZE):
            batch = rows[i : i + BATCH_SIZE]
            scripts = [
                Script(
                    id=str(r["Script_ID"]).strip(),
                    title=str(r["Script_Title"]).strip(),
                    purpose=str_or_none(r.get("Script_Purpose")),
                    module=str_or_none(r.get("Module")),
                    category=str_or_none(r.get("Category")),
                    script_text=str(r["Script_Text_Sanitized"]).strip(),
                )
                for r in batch
            ]
            session.add_all(scripts)
            await session.flush()
        print(f"  Scripts: {len(rows)}")

        # 4. Tickets (FK → knowledge_articles, scripts)
        rows = read_sheet(wb, "Tickets")
        for i in range(0, len(rows), BATCH_SIZE):
            batch = rows[i : i + BATCH_SIZE]
            tickets = [
                Ticket(
                    id=str(r["Ticket_Number"]).strip(),
                    status=str(r.get("Status", "open")).strip().lower(),
                    priority=str(r.get("Priority", "medium")).strip().lower(),
                    product=str_or_none(r.get("Product")),
                    module=str_or_none(r.get("Module")),
                    category=str_or_none(r.get("Category")),
                    description=str_or_none(r.get("Description")),
                    resolution=str_or_none(r.get("Resolution")),
                    root_cause=str_or_none(r.get("Root_Cause")),
                    tags=str_or_none(r.get("Tags")),
                    kb_article_id=str_or_none(r.get("KB_Article_ID")),
                    script_id=str_or_none(r.get("Script_ID")),
                )
                for r in batch
            ]
            session.add_all(tickets)
            await session.flush()
        print(f"  Tickets: {len(rows)}")

        # 5. Conversations (FK → tickets)
        rows = read_sheet(wb, "Conversations")
        for i in range(0, len(rows), BATCH_SIZE):
            batch = rows[i : i + BATCH_SIZE]
            conversations = [
                Conversation(
                    id=str(r["Conversation_ID"]).strip(),
                    ticket_id=str(r["Ticket_Number"]).strip(),
                    channel=str_or_none(r.get("Channel")),
                    agent_name=str_or_none(r.get("Agent_Name")),
                    transcript=str_or_none(r.get("Transcript")),
                    sentiment=str_or_none(r.get("Sentiment")),
                    conversation_start=parse_datetime(r.get("Conversation_Start")),
                    conversation_end=parse_datetime(r.get("Conversation_End")),
                )
                for r in batch
            ]
            session.add_all(conversations)
            await session.flush()
        print(f"  Conversations: {len(rows)}")

        # 6. Questions (polymorphic target_id, no FK)
        rows = read_sheet(wb, "Questions")
        for i in range(0, len(rows), BATCH_SIZE):
            batch = rows[i : i + BATCH_SIZE]
            questions = [
                Question(
                    id=str(r["Question_ID"]).strip(),
                    source=str_or_none(r.get("Source")),
                    product=str_or_none(r.get("Product")),
                    category=str_or_none(r.get("Category")),
                    module=str_or_none(r.get("Module")),
                    difficulty=str_or_none(r.get("Difficulty")),
                    question_text=str(r["Question_Text"]).strip(),
                    answer_type=str_or_none(r.get("Answer_Type")),
                    target_id=str_or_none(r.get("Target_ID")),
                )
                for r in batch
            ]
            session.add_all(questions)
            await session.flush()
        print(f"  Questions: {len(rows)}")

        # 7. KB Lineage (FK → knowledge_articles)
        rows = read_sheet(wb, "KB_Lineage")
        lineage_items = [
            KBLineage(
                kb_article_id=str(r["KB_Article_ID"]).strip(),
                source_id=str_or_none(r.get("Source_ID")),
                relationship=str_or_none(r.get("Relationship")),
                evidence_snippet=str_or_none(r.get("Evidence_Snippet")),
                event_timestamp=parse_datetime(r.get("Event_Timestamp")),
            )
            for r in rows
        ]
        session.add_all(lineage_items)
        await session.flush()
        print(f"  KB Lineage: {len(lineage_items)}")

        # 8. Learning Events (FK → tickets, knowledge_articles)
        rows = read_sheet(wb, "Learning_Events")
        events = [
            LearningEvent(
                id=str(r["Event_ID"]).strip(),
                trigger_ticket_id=str_or_none(r.get("Trigger_Ticket_Number")),
                detected_gap=str_or_none(r.get("Detected_Gap")),
                proposed_kb_article_id=str_or_none(r.get("Proposed_KB_Article_ID")),
                final_status=str_or_none(r.get("Final_Status")),
            )
            for r in rows
        ]
        session.add_all(events)
        await session.flush()
        print(f"  Learning Events: {len(events)}")

        await session.commit()
        print("Import complete.")

    wb.close()
    await engine.dispose()


def main() -> None:
    asyncio.run(import_all())


if __name__ == "__main__":
    main()
